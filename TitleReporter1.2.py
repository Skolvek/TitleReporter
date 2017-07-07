#!Python 3
#TitleReporter1.2.py
#Stephen Kolvek JULY 6, 2017
#New with ver1.2: updated UI, fixed bugs

import os, openpyxl, sys
import openpyxl.cell
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

#define each bank
banks = ['BRISTOL', 'COMMERCE', 'EASTERN', 'GREENWOOD',
         'JJBEST', 'NAVIGANT', 'SIGNATURE', 'VIRGINIA']

#LOAD THE FILES FROM DIRECTORY (Enter the directory you are working in here)
os.chdir(r'C:\Users\Stephen\Documents\J.J. Best Work Stuff\JUNE OUTSTANDING TITLES\WORKABLE_XLSX_FILES')

#Delete the column containing currency values (WORKING)
def delete_column(col):
    for rowNum in range(1, sheet.max_row + 1):
        sheet.cell(row=rowNum, column=col).value = None

#Format Col as 'short date' (WORKING)
def format_col_as_date(col):
    for i in range(2, sheet.max_row + 1):
        sheet.cell(row=i, column=col).number_format = 'MM/DD/YYYY'

#Set width of Cols A-L (WORKING)
def set_col_widths():

    #define widths for each column
    widths = [20, 16.50, 12, 8, 13, 6,
              15, 15, 15, 26, 12, 80]
    #iterate through and set 'em
    for i in range(1, sheet.max_column):
        sheet.column_dimensions[get_column_letter(i)].width = widths[i-1]
            
#Text Wrapping on cols I, J, L (WORKING)
def wrap_text_in_col(num):
    for r in range(2, sheet.max_row + 1):
        sheet.cell(row=r, column=num).alignment = sheet.cell(row=r, column=num).alignment.copy(wrapText=True)

#set cells to 'top align' (WORKING)
def top_align_cells():
    #Iterate in the 2D array of the sheet skipping top row and far right col (Thus the indexes) 
    for j in range(1, sheet.max_column-1):
        for k in range(2, sheet.max_row + 1):
            #Set the alignment for cells in the given range above k is row; j is col
            sheet.cell(row=k, column=j).alignment = sheet.cell(row=k, column=j).alignment.copy(horizontal='left', vertical='top')

#Set header names (WORKING)
def set_headers():
    
    #cleaned this up 7/1/17
    headers = ['Customer','Status','Date','State',
               'Ammount','Year','Make','Model','Owner',
               'Seller','Proof Date','Title Notes']

    for i in range(1, sheet.max_column):
        sheet.cell(row=1, column=i).value = headers[i-1]

#7/6/17 more streamlined UI

print('Welcome to TitleReporterV1.2 written by Stephen Kolvek')
print('For this program to work, name the raw files like this:'
      '\n<BANKNAME>_OTR_ROUGH.xlsx')

while True:
   
    print('If you are ready type \'r\' to run or \'q\' to quit.')
    choice = input()
    try:
        
        if choice == 'r' or choice == 'R':

            for i in range(len(banks)):
                otr = openpyxl.load_workbook(banks[i]+'_OTR_ROUGH.xlsx')
                sheet = otr.get_active_sheet()
                set_headers()
                set_col_widths()
                delete_column(13)

                for k in (9, 10, 12):
                    wrap_text_in_col(k)

                for j in (3, 11):
                    format_col_as_date(j)

                top_align_cells()
                #tell the user the file name
                print('CREATED: FINAL_REPORT_'+banks[i]+'.xlsx') 
                otr.save('FINAL_REPORT_'+banks[i]+'.xlsx')
        
        elif choice == 'q' or choice == 'Q':
            print('Till next time!')
            sys.exit()
    
    except FileNotFoundError:
        print('ERROR: File Not Found. Check that all Files were in the '
              '\n current working directory')
        print()

